import os
import openpyxl
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import concurrent.futures
import threading
import psutil
import win32file
import win32com.client
import queue
import pandas as pd


class ThreadSafeLogger:
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.log_queue = queue.Queue()
        self.stop_event = threading.Event()
        self.log_thread = threading.Thread(target=self._log_worker, daemon=True)
        self.log_thread.start()

    def _log_worker(self):
        while not self.stop_event.is_set():
            try:
                message = self.log_queue.get(timeout=0.1)
                self.text_widget.insert(tk.END, message + "\n")
                self.text_widget.see(tk.END)
            except queue.Empty:
                continue

    def log(self, message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_queue.put(log_message)

    def stop(self):
        self.stop_event.set()
        self.log_thread.join()


class ExcelProcessorApp:
    def __init__(self, master):
        self.master = master
        master.title("Excel Processor")
        master.geometry("800x600")

        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Processing Tab
        self.processing_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.processing_frame, text="File Processing")

        # Details Tab
        self.details_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.details_frame, text="Processed Details")

        # Processed files storage
        self.processed_files = []

        # Setup Processing Tab
        self.setup_processing_tab()

        # Setup Details Tab
        self.setup_details_tab()

    def setup_processing_tab(self):
        # Directory path label and entry
        dir_label = ttk.Label(self.processing_frame, text="Excel Files Directory:")
        dir_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)

        self.dir_entry = ttk.Entry(self.processing_frame, width=40)
        self.dir_entry.grid(row=0, column=1, padx=5, pady=5)
        self.dir_entry.insert(0, r"M:\S A L E S\SHOWROOM\QUOTATION\2024")

        # Browse button
        browse_button = ttk.Button(self.processing_frame, text="Browse", command=self.browse_directory)
        browse_button.grid(row=0, column=2, padx=5, pady=5)

        # Process button
        self.process_button = ttk.Button(self.processing_frame, text="Process Files",
                                         command=self.start_file_processing)
        self.process_button.grid(row=1, column=2, padx=5, pady=5)

        # Log text area
        self.log_text = tk.Text(self.processing_frame, height=10, width=60, state=tk.NORMAL)
        self.log_text.grid(row=2, column=0, columnspan=3, padx=5, pady=5, sticky='nsew')

        # Scrollbar for log text
        log_scrollbar = ttk.Scrollbar(self.processing_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scrollbar.grid(row=2, column=3, sticky='ns')
        self.log_text.configure(yscroll=log_scrollbar.set)

        # Progress bar
        self.progress_bar = ttk.Progressbar(self.processing_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.grid(row=3, column=1, padx=5, pady=5)

        # Configure grid
        self.processing_frame.grid_columnconfigure(1, weight=1)
        self.processing_frame.grid_rowconfigure(2, weight=1)

        # Setup thread-safe logger
        self.logger = ThreadSafeLogger(self.log_text)
        self.logger.log("Excel Processor App started.")

    def setup_details_tab(self):
        # Treeview to display processed file details
        self.details_tree = ttk.Treeview(self.details_frame, columns=(
            "Filename", "Date", "Quotation No", "Prepared By",
            "Customer Details", "Contact Name", "Contact", "Fax No", "Email", "TIN No"
        ), show='headings')

        # Define headings
        for col in self.details_tree['columns']:
            self.details_tree.heading(col, text=col)
            self.details_tree.column(col, width=100)

        self.details_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Scrollbar for treeview
        details_scrollbar = ttk.Scrollbar(self.details_frame, orient=tk.VERTICAL, command=self.details_tree.yview)
        details_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.details_tree.configure(yscroll=details_scrollbar.set)

        # Buttons for details tab
        button_frame = ttk.Frame(self.details_frame)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        export_csv_button = ttk.Button(button_frame, text="Export to CSV", command=self.export_to_csv)
        export_csv_button.pack(side=tk.LEFT, padx=5)

        export_excel_button = ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel)
        export_excel_button.pack(side=tk.LEFT, padx=5)

        clear_button = ttk.Button(button_frame, text="Clear Details", command=self.clear_details)
        clear_button.pack(side=tk.LEFT, padx=5)

    def browse_directory(self):
        """Open directory browser to select Excel files directory"""
        directory = filedialog.askdirectory()
        if directory:
            self.dir_entry.delete(0, tk.END)
            self.dir_entry.insert(0, directory)

    def safe_load_workbook(self, file_path):
        """Safely load an Excel workbook"""
        try:
            # Check file accessibility
            win32file.CreateFile(
                file_path,
                win32file.GENERIC_READ,
                win32file.FILE_SHARE_READ | win32file.FILE_SHARE_WRITE,
                None,
                win32file.OPEN_EXISTING,
                win32file.FILE_ATTRIBUTE_NORMAL,
                None
            )

            # Load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            return wb
        except Exception as e:
            self.logger.log(f"Cannot access file {file_path}: {e}")
            return None

    def extract_excel_details(self, file_path):
        """Extract details from Excel file"""
        try:
            # Load workbook safely
            wb = self.safe_load_workbook(file_path)
            if not wb:
                return None

            # Select the first worksheet
            ws = wb.active

            # Extract details
            details = {
                'filename': os.path.basename(file_path),
                'date': str(ws['J9'].value) if ws['J9'].value else '',
                'quotation_no': str(ws['J10'].value) if ws['J10'].value else '',
                'prepared_by': str(ws['J11'].value) if ws['J11'].value else '',
                'tin_no': str(ws['J12'].value) if ws['J12'].value else '',
                'customer_detail_1': str(ws['A15'].value) if ws['A15'].value else '',
                'customer_detail_2': str(ws['A16'].value) if ws['A16'].value else '',
                'customer_detail_3': str(ws['A17'].value) if ws['A17'].value else '',
                'customer_detail_4': str(ws['A18'].value) if ws['A18'].value else '',
                'customer_detail_5': str(ws['A19'].value) if ws['A19'].value else '',
                'customer_contact_name': str(ws['J14'].value) if ws['J14'].value else '',
                'customer_contact': str(ws['J15'].value) if ws['J15'].value else '',
                'customer_fax_no': str(ws['J16'].value) if ws['J16'].value else '',
                'customer_email': str(ws['J17'].value) if ws['J17'].value else '',
                'customer_tin_no': str(ws['J18'].value) if ws['J18'].value else '',
            }

            return details

        except Exception as e:
            self.logger.log(f"Error processing {os.path.basename(file_path)}: {e}")
            return None

    def process_single_file(self, file_path):
        """Process a single file"""
        try:
            self.logger.log(f"Processing {os.path.basename(file_path)}")
            file_details = self.extract_excel_details(file_path)

            if file_details:
                self.display_details(file_details)
                self.processed_files.append(file_details)  # Store processed details

            return file_details
        except Exception as e:
            self.logger.log(f"Unexpected error processing {file_path}: {e}")
            return None

    def start_file_processing(self):
        """Start file processing in a separate thread"""

        def process_files_thread():
            try:
                # Disable process button during processing
                self.process_button.config(state=tk.DISABLED)

                # Get the directory path
                directory_path = self.dir_entry.get()

                # Supported Excel file extensions
                excel_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']

                # Get list of Excel files
                excel_files = [
                    os.path.join(directory_path, filename)
                    for filename in os.listdir(directory_path)
                    if any(filename.lower().endswith(ext) for ext in excel_extensions)
                ]

                # Determine optimal thread count
                max_workers = min(len(excel_files), max(psutil.cpu_count() - 1, 1))

                # Start multithreaded processing
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = [executor.submit(self.process_single_file, file_path) for file_path in excel_files]

                    # Update progress bar
                    total_files = len(futures)
                    for i, future in enumerate(concurrent.futures.as_completed(futures), 1):
                        progress = int((i / total_files) * 100)
                        self.master.after(0, self.update_progress, progress)

                self.logger.log("File processing complete.")
            finally:
                # Re-enable process button after processing
                self.process_button.config(state=tk.NORMAL)

        # Start the file processing in a new thread
        threading.Thread(target=process_files_thread, daemon=True).start()

    def update_progress(self, value):
        """Update the progress bar"""
        self.progress_bar['value'] = value

    def display_details(self, details):
        """Display extracted details in the log and details tab"""
        self.logger.log(f"\nFile: {details['filename']}")
        self.logger.log(f"Date: {details['date']}")
        self.logger.log(f"Quotation No: {details['quotation_no']}")
        self.logger.log("-" * 40)

        # Combine customer details into a single string
        customer_details = f"{details.get('customer_detail_1', '')} | {details.get('customer_detail_2', '')} | {details.get('customer_detail_3', '')} | {details.get('customer_detail_4', '')} | {details.get('customer_detail_5', '')}".strip()

        # Insert details into the treeview
        self.details_tree.insert("", "end", values=(
            details['filename'],
            details['date'],
            details['quotation_no'],
            details['prepared_by'],
            customer_details,
            details.get('customer_contact_name', ''),
            details['customer_contact'],
            details['customer_fax_no'],
            details['customer_email'],
            details['customer_tin_no']
        ))

    def export_to_csv(self):
        """Export processed details to a CSV file"""
        if not self.processed_files:
            messagebox.showwarning("No Data", "No processed details to export.")
            return

        df = pd.DataFrame(self.processed_files )
        file_path = filedialog.asksaveasfilename(defaultextension=".csv",
                                                 filetypes=[("CSV files", "*.csv")])
        if file_path:
            df.to_csv(file_path, index=False)
            messagebox.showinfo("Export Successful", "Processed details exported to CSV.")

    def export_to_excel(self):
        """Export processed details to an Excel file"""
        if not self.processed_files:
            messagebox.showwarning("No Data", "No processed details to export.")
            return

        df = pd.DataFrame(self.processed_files)
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Export Successful", "Processed details exported to Excel.")

    def clear_details(self):
        """Clear the details displayed in the treeview"""
        self.details_tree.delete(*self.details_tree.get_children())
        self.processed_files.clear()
        self.logger.log("Processed details cleared.")


root = tk.Tk()
app = ExcelProcessorApp(root)
root.mainloop()